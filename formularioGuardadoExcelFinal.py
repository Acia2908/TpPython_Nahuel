import tkinter as tk
from openpyxl import Workbook

# Crear la ventana principal
ventana = tk.Tk()

#root=tk()
#tk.title("Entrada de datos")
#ventana.geometry("500x300")

ventana.title("Formulario Ingreso Datos")

def guardar_en_excel():
    
    datos = []
    for row in range(3):
        row_datos = []
        for col in range(num_columnas.get()):
            valor = cajas_texto[row][col].get()
            row_datos.append(valor)
        datos.append(row_datos)

    # Crear un nuevo archivo de Excel
    workbook = Workbook()
    sheet = workbook.active

    # Escribir los datos en las celdas a parir de la segunda
    for row_num, row_data in enumerate(datos, start=1):
        for col_num, value in enumerate(row_data, start=1):
            sheet.cell(row=1, column=1, value="Apellido")# DEJE fijas las columnas de cabecera .. 
            sheet.cell(row=2, column=1, value="Nombre")
            sheet.cell(row=3, column=1, value="Telefono")
            sheet.cell(row=row_num, column=col_num+1, value=value)

    # Guardar el archivo .. mismo directorio ... le saque la ruta
    workbook.save("inventarioDatosCliente.xlsx")
    #muestro por consola como se llama el archivo que genere.. 
    # se me pisa la informacion ... no logro que sea acumulativa ...
    print("Datos guardados en el archivo 'inventarioDatosCliente.xlsx'.")


# Crear el campo de entrada para el número de columnas
num_columnas_label = tk.Label(ventana, text="Número de clientes a cargar :")
num_columnas_label.grid(row=1, column=0)
num_columnas = tk.IntVar()
num_columnas_entry = tk.Entry(ventana, textvariable=num_columnas)
num_columnas_entry.grid(row=1, column=1)

# Obtener los datos ingresados en las cajas de texto
#seccion de Nombre
nombre_label= tk.Label(ventana, text="Ingresar el Apellido:")
nombre_label.grid(row=2, column=0)
nombre_label= tk.Label(ventana, text="Ingresar el Nombre:")
nombre_label.grid(row=3, column=0)
#nombre_label.config(padx=10, pady=10)
#seccion de Telefono
nombre_label= tk.Label(ventana, text="Ingresar el Telefono:")
nombre_label.grid(row=4, column=0)


# Función para crear las cajas de texto
def crear_cajas_texto():
  
    num_cols = num_columnas.get()
    for row in range(3):
        for col in range(num_cols):
            caja_texto = tk.Entry(ventana)
            caja_texto.grid(row=row+2, column=col+1)# le sume valores para que inicien la creacion de las cajitas luego de la primer columna de texto ... 
            cajas_texto[row].append(caja_texto)
            caja_texto.delete(0, tk.END)
 #FRAME DE ENTRADA DE DATOS

bienvenido = tk.Label(ventana, text="BIENVENIDOS")
bienvenido.grid(row=0, column=0, columnspan = 2)
bienvenido.config(font=('Arial', 16))

# Crear el botón para crear las cajas de texto lo deje fijo ...arriba asi no queda detras del area a ingresar datos
crear_button = tk.Button(ventana, text="Crear lugar de ingreso de data", command=crear_cajas_texto)
crear_button.grid(row=1, column=2, columnspan = 2)

# Crear las cajas de texto iniciales
cajas_texto = [[] for _ in range(3)]

# Crear el botón para guardar los datos
guardar_button = tk.Button(ventana, text="Guardar-Archivo", command=guardar_en_excel)
guardar_button.grid(row=5, column=2, columnspan=2)

# Iniciar el bucle principal de la ventana
ventana.mainloop()
